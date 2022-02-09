import pickle
import openpyxl
from openpyxl import load_workbook

# We have already scraped all the data and html pages we need. 
# Now we need to parse through pkl files and save the items into excel spreadsheets

# pkl > array > 

print('opening chin1')
chin1 = pickle.load(open('soups/individual_pages/Chin1.pkl', 'rb'))
print('opening chin2')
chin2 = pickle.load(open('soups/individual_pages/Chin2.pkl', 'rb'))
print('opening ching')
ching = pickle.load(open('soups/individual_pages/Ching.pkl', 'rb'))
print('opening dea')
dea = pickle.load(open('soups/individual_pages/Dea.pkl', 'rb'))
print('opening der')
der = pickle.load(open('soups/individual_pages/Der.pkl', 'rb'))
print('opening fong1')
fong1 = pickle.load(open('soups/individual_pages/Fong1.pkl', 'rb'))
print('opening fong2')
fong2 = pickle.load(open('soups/individual_pages/Fong2.pkl', 'rb'))
print('opening hong')
hong = pickle.load(open('soups/individual_pages/Hong.pkl', 'rb'))
print('opening jair')
jair = pickle.load(open('soups/individual_pages/Jair.pkl', 'rb'))
print('opening tan')
tan = pickle.load(open('soups/individual_pages/Tan.pkl', 'rb'))
print('opening tsing')
tsing = pickle.load(open('soups/individual_pages/Tsing.pkl', 'rb'))

individual_pages = [chin1, chin2, ching, dea, der, fong1, fong2, hong, 
jair, tan, tsing]


# individual_pages = [chin1, chin2]

individual_names = ['chin1', 'chin2', 'ching', 'dea', 'der', 'fong1', 'fong2', 'hong', 'jair', 'tan', 'tsing']


def single_individual(html):
    attributes, values = [], []
    tds = html.find_all('td')
    for i in range(len(tds)):
        if i%2:
            values.append(tds[i].text)
        else:
            attributes.append(tds[i].text)
    # for i in range(len(attributes)):
    #     print(attributes[i], ' = ', values[i])
    return [attributes, values]


# we create the spreadsheet
def create_spreadsheet():
    filepath = "final_spreadsheet.xlsx"
    wb = openpyxl.Workbook()
    wb.save(filepath)


# ws['A1'] = 1
# ws.cell(row=2, column=2).value = 2
# wb.save('final_spreadsheet.xlsx')


# We need to add single_individual to the spreadsheet
def values_to_spreadsheet(attributes, values, row_n):
    wb = load_workbook(filename='final_spreadsheet.xlsx')
    ws = wb.worksheets[0]
    # row_n = 2
    # cell = ws.cell(row = row_n, column = 1).value
    # while cell is not None:
    #     row_n += 1
    #     cell = ws.cell(row = row_n, column = 1).value
    for i in range(len(attributes)):
        if ws.cell(row = 1, column = i + 1).value is None:
            ws.cell(row = 1, column = i + 1).value = attributes[i]
        ws.cell(row = row_n, column = i + 1).value = values[i]
    wb.save('final_spreadsheet.xlsx')


def first_empty_row():
    wb = load_workbook(filename='final_spreadsheet.xlsx')
    ws = wb.worksheets[0]
    row_n = 2
    cell = ws.cell(row = row_n, column = 1).value
    cell2 = ws.cell(row = row_n, column = 2).value
    cell4 = ws.cell(row = row_n, column = 4).value
    while cell is not None or cell2 is not None or cell4 is not None:
        row_n += 1
        cell = ws.cell(row = row_n, column = 1).value
        cell2 = ws.cell(row = row_n, column = 2).value
        cell4 = ws.cell(row = row_n, column = 4).value
    return row_n

def add_everything_to_spreadsheet():
    # create_spreadsheet()
    row = first_empty_row()
    counter = 2
    for i in range(len(individual_pages)):
        for j in range(len(individual_pages[i])):
            if counter > row:
                print('adding ', individual_names[i], ' row ', row)
                attributes = single_individual(individual_pages[i][j])[0]
                values = single_individual(individual_pages[i][j])[1]
                values_to_spreadsheet(attributes, values, row)
                row += 1
            else:
                print('skipped ', counter)
            counter += 1


# html = single_individual(chin1[2])

# attributes = single_individual(chin1[2])[0]
# values = single_individual(chin1[2])[1]
# values_to_spreadsheet(attributes, values)


add_everything_to_spreadsheet()
# wb = load_workbook('final_spreadsheet.xlsx')